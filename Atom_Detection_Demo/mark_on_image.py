import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
# Image width and height (nm)
image_width = 16.411
image_height = 16.411

current_directory = os.getcwd()
path_in = os.path.join(current_directory, "data", "tif_data")
path_csv = os.path.join(current_directory, "data", "detection_data")
path_out = os.path.join(current_directory, "data", "tif_data", "result")

if not os.path.exists(path_out):
    os.makedirs(path_out)


def get_files_with_extensions(folder_path):
    file_list = []
    dir_list = []
    for root, directories, files in sorted(os.walk(folder_path)):
        for file in files:
            file_list.append(os.path.join(root, file))
        for dir in directories:
            dir_list.append(dir)
    return file_list, dir_list


files_with_extensions, directories = get_files_with_extensions(path_csv)
ii = 0

# Print the file list
for file in sorted(files_with_extensions):
    ii += 1
    base = os.path.splitext(os.path.basename(file))[0]

    fni = os.path.join(path_in, base + '.tif')
    fnout = os.path.join(path_out, base + '_out.tif')

    x = []
    y = []
    img = plt.imread(fni)  # your image
    size = 50

    df1 = pd.read_csv(file)

    sheet.cell(row=ii, column=1).value = base
    sheet.cell(row=ii, column=2).value = len(df1.index) / image_width / image_height

    ax = df1.plot(kind='scatter', x='y', y='x', marker='.', s=10, color='cyan')
    plt.axis('off')
    plt.imshow(img, cmap="gray")  # plot image
    plt.savefig(fnout, bbox_inches='tight', pad_inches=0.0, dpi=300)
    plt.close()

    # Save Excel workbook inside the loop
    base_dir = os.path.join(path_out, base)

    if not os.path.exists(base_dir):
        os.makedirs(base_dir)

    wb.save(os.path.join(base_dir, 'Density.xlsx'))
