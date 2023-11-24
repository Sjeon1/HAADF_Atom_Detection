import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
# Image width and height (nm)
image_width = 16.411
image_height = 16.411

current_directory = os.getcwd()
path_in = os.path.join(current_directory, "data", "tif_data")
path_csv = os.path.join(current_directory, "data", "detection_data", "dl_detection_ReBubpy", "dl_detection_ReBubpy_0.6")
path_out = os.path.join(current_directory, "data", "tif_data", "result")

if not os.path.exists(path_out):
   os.makedirs(path_out)

def get_files_with_extensions(folder_path):
    file_list = []
    for root, directories, files in sorted(os.walk(folder_path)):
        for file in files:
            file_list.append(file)
    return file_list

files_with_extensions = get_files_with_extensions(path_csv)
ii = 0

# Print the file list
for file in sorted(files_with_extensions):
    ii += 1
    base = os.path.splitext(file)[0]
    
    fni = base + '.tif'
    fncsv = base + '.csv'
    fnout = base + 'out' + '.tif'

    x = []
    y = []
    img = plt.imread(os.path.join(path_in, fni))  # your image
    size = 50
    
    df1 = pd.read_csv(os.path.join(path_csv, fncsv))
    
    sheet.cell(row=ii, column=1).value = base
    sheet.cell(row=ii, column=2).value = len(df1.index) / image_width / image_height
    
    ax = df1.plot(kind='scatter', x='y', y='x', marker='.', s=10, color='cyan')
    plt.axis('off')
    plt.imshow(img, cmap="gray")  # plot image
    plt.savefig(os.path.join(path_out, fnout), bbox_inches='tight', pad_inches=0.0, dpi=300)
    plt.close()

# Save Excel workbook outside the loop
wb.save(os.path.join(path_out, 'Density.xlsx'))

